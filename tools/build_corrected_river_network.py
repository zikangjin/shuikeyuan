from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import geopandas as gpd
import pandas as pd
import yaml
from openpyxl import load_workbook
from shapely.geometry import LineString, Point


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.export_utils import safe_filename  # noqa: E402


RIVER_WORDS = ("河中", "河道", "河底", "河槽", "沟", "渠", "水沟")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据所有断面汇总表的河中点生成校正后河道。")
    parser.add_argument("--config", default="config.yaml", help="配置文件路径，默认 config.yaml")
    parser.add_argument("--section-root", default=None, help="断面数据根目录；不填则读取 config.yaml 的 section_path")
    parser.add_argument("--output-dir", default=None, help="输出目录；不填则写入 config.yaml 的 output_dir/corrected_rivers")
    parser.add_argument("--crs", default="EPSG:4548", help="断面河中点坐标系，默认 EPSG:4548")
    return parser.parse_args()


def load_config(path: Path) -> dict:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as file:
        return yaml.safe_load(file) or {}


def split_paths(value) -> list[Path]:
    if value in (None, ""):
        return []
    if isinstance(value, (list, tuple)):
        return [Path(str(item)) for item in value if str(item).strip()]
    return [Path(item.strip()) for item in re.split(r"[;\n]+", str(value)) if item.strip()]


def find_excel_tables(paths: list[Path]) -> list[Path]:
    tables: list[Path] = []
    for path in paths:
        if path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$"):
            tables.append(path)
            continue
        if not path.exists() or not path.is_dir():
            continue
        for item in path.rglob("*.xlsx"):
            if item.name.startswith("~$"):
                continue
            name = item.name
            if any(word in name for word in ("汇总", "断面数据", "横断", "断面")):
                tables.append(item)
    seen = set()
    unique: list[Path] = []
    for table in tables:
        key = str(table.resolve()).lower() if table.exists() else str(table).lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(table)
    return unique


def extract_xy(row) -> tuple[float | None, float | None]:
    nums: list[float] = []
    for value in row:
        if value is None:
            continue
        try:
            nums.append(float(value))
        except (TypeError, ValueError):
            continue
    for idx in range(len(nums) - 1):
        first = nums[idx]
        second = nums[idx + 1]
        if 300000 <= first <= 700000 and 4300000 <= second <= 4600000:
            return first, second
        if 4300000 <= first <= 4600000 and 300000 <= second <= 700000:
            return second, first
    xs = [value for value in nums if 300000 <= value <= 700000]
    ys = [value for value in nums if 4300000 <= value <= 4600000]
    if xs and ys:
        return xs[0], ys[0]
    return None, None


def section_order(value: str) -> int:
    numbers = re.findall(r"\d+", str(value or ""))
    return int(numbers[-1]) if numbers else 0


def short_river_name(folder_name: str) -> str:
    text = str(folder_name or "").strip()
    text = re.split(r"[-_]?20\d{2}|坐标|国家", text)[0]
    return text.strip("-_ ") or str(folder_name or "未命名河道")


def row_to_record(row, table_path: Path, sheet_name: str) -> dict | None:
    if not row or len(row) < 5:
        return None
    try:
        section_name = str(row[1]).strip()
        note = "" if row[2] is None else str(row[2]).strip()
        station = float(row[3])
        h85 = float(row[4]) if row[4] is not None else None
    except Exception:
        return None
    if not section_name or section_name.lower() == "none":
        return None
    if "纵断" in section_name or "纵断" in str(sheet_name):
        return None
    x, y = extract_xy(row)
    if x is None or y is None:
        return None
    return {
        "river_folder": table_path.parent.name,
        "river_name": short_river_name(table_path.parent.name),
        "section_name": section_name,
        "section_order": section_order(section_name),
        "note": note,
        "station_m": station,
        "h85_m": h85,
        "x": x,
        "y": y,
        "table_path": str(table_path),
        "sheet_name": sheet_name,
    }


def choose_center_record(records: list[dict]) -> dict | None:
    valid = [item for item in records if item.get("x") is not None and item.get("y") is not None]
    if not valid:
        return None

    center = [item for item in valid if "河中" in str(item.get("note", ""))]
    if center:
        center = sorted(center, key=lambda item: item["station_m"])
        item = center[len(center) // 2].copy()
        item["center_source"] = "河中"
        return item

    river = [item for item in valid if any(word in str(item.get("note", "")) for word in RIVER_WORDS)]
    if river:
        item = min(river, key=lambda row: row["h85_m"] if row.get("h85_m") is not None else float("inf")).copy()
        item["center_source"] = "河道最低点"
        return item

    with_h85 = [item for item in valid if item.get("h85_m") is not None]
    if with_h85:
        item = min(with_h85, key=lambda row: row["h85_m"]).copy()
        item["center_source"] = "断面最低点"
        return item

    item = valid[len(valid) // 2].copy()
    item["center_source"] = "断面中间点"
    return item


def collect_center_points(tables: list[Path]) -> tuple[list[dict], list[dict]]:
    centers: list[dict] = []
    diagnostics: list[dict] = []
    for table_path in tables:
        try:
            workbook = load_workbook(table_path, data_only=True, read_only=True)
        except Exception as exc:
            diagnostics.append({"table_path": str(table_path), "status": "读取失败", "note": str(exc)})
            continue

        by_section: dict[tuple[str, str], list[dict]] = {}
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                record = row_to_record(row, table_path, worksheet.title)
                if record is None:
                    continue
                key = (worksheet.title, record["section_name"])
                by_section.setdefault(key, []).append(record)

        table_count = 0
        for (_sheet_name, section_name), records in by_section.items():
            center = choose_center_record(records)
            if center is None:
                continue
            center["point_count_in_section"] = len(records)
            centers.append(center)
            table_count += 1
        diagnostics.append(
            {
                "table_path": str(table_path),
                "river_folder": table_path.parent.name,
                "status": "成功",
                "center_count": table_count,
            }
        )
    return centers, diagnostics


def build_outputs(centers: list[dict], output_dir: Path, crs: str) -> tuple[Path, Path, Path, pd.DataFrame]:
    output_dir.mkdir(parents=True, exist_ok=True)
    centers_df = pd.DataFrame(centers)
    if centers_df.empty:
        raise ValueError("没有从断面汇总表中提取到任何可用于校正的河中点。")

    points_gdf = gpd.GeoDataFrame(
        centers_df.copy(),
        geometry=[Point(xy) for xy in zip(centers_df["x"], centers_df["y"])],
        crs=crs,
    )

    line_rows = []
    for river_folder, group in centers_df.groupby("river_folder", sort=False):
        group = group.sort_values(["section_order", "section_name", "station_m"])
        coords: list[tuple[float, float]] = []
        section_names: list[str] = []
        center_sources: list[str] = []
        for _, row in group.iterrows():
            xy = (float(row["x"]), float(row["y"]))
            if not coords or xy != coords[-1]:
                coords.append(xy)
            section_names.append(str(row["section_name"]))
            center_sources.append(str(row.get("center_source", "")))
        if len(coords) < 2:
            continue
        line_rows.append(
            {
                "river_folder": river_folder,
                "river_name": short_river_name(river_folder),
                "section_count": len(section_names),
                "center_sources": "、".join(sorted(set(center_sources))),
                "sections": "、".join(section_names[:80]) + (f" ... 共 {len(section_names)} 个" if len(section_names) > 80 else ""),
                "geometry": LineString(coords),
            }
        )

    lines_gdf = gpd.GeoDataFrame(line_rows, geometry="geometry", crs=crs)
    if lines_gdf.empty:
        raise ValueError("河中点数量不足，无法生成线状校正河道。")

    line_path = output_dir / "校正后河道_全部断面.geojson"
    point_path = output_dir / "校正河中点_全部断面.geojson"
    excel_path = output_dir / "校正河道诊断_全部断面.xlsx"

    lines_gdf.to_file(line_path, driver="GeoJSON", encoding="utf-8")
    points_gdf.to_file(point_path, driver="GeoJSON", encoding="utf-8")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        centers_df.to_excel(writer, sheet_name="river_center_points", index=False)
        pd.DataFrame(line_rows).drop(columns=["geometry"], errors="ignore").to_excel(writer, sheet_name="corrected_rivers", index=False)

    return line_path, point_path, excel_path, centers_df


def main() -> int:
    args = parse_args()
    config_path = Path(args.config)
    config = load_config(config_path)

    section_paths = split_paths(args.section_root) if args.section_root else split_paths(config.get("section_paths") or config.get("section_path"))
    if not section_paths:
        raise SystemExit("没有指定断面数据目录。请使用 --section-root 或在 config.yaml 中填写 section_path。")

    output_dir = Path(args.output_dir) if args.output_dir else Path(config.get("output_dir") or "outputs") / "corrected_rivers"
    tables = find_excel_tables(section_paths)
    if not tables:
        raise SystemExit(f"未找到断面 Excel 汇总表：{'; '.join(str(path) for path in section_paths)}")

    centers, diagnostics = collect_center_points(tables)
    line_path, point_path, excel_path, centers_df = build_outputs(centers, output_dir, args.crs)

    diagnostics_path = output_dir / "断面表读取诊断.xlsx"
    pd.DataFrame(diagnostics).to_excel(diagnostics_path, index=False)

    river_count = centers_df["river_folder"].nunique()
    print(f"读取断面表：{len(tables)} 个")
    print(f"提取河中/河道中心点：{len(centers_df)} 个，涉及河道：{river_count} 条")
    print(f"校正后河道：{line_path}")
    print(f"校正河中点：{point_path}")
    print(f"诊断表：{excel_path}")
    print(f"读取诊断：{diagnostics_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
