from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import geopandas as gpd
import pandas as pd
import rasterio
from openpyxl import load_workbook
from rasterio.crs import CRS
from shapely import wkt
from shapely.geometry import LineString, Point


VILLAGE_FIELD_CANDIDATES = [
    "village",
    "村名",
    "村庄名称",
    "行政村",
    "Name",
    "name",
    "NAME",
    "XZQMC",
    "CUNMC",
    "地名",
]
BUILDING_ID_CANDIDATES = ["OBJECTID", "FID", "id", "ID", "Id", "建筑物编号", "建筑物名称", "name", "Name"]
SECTION_ID_CANDIDATES = ["section_id", "断面编号", "断面号", "编号", "id", "ID", "FID", "OBJECTID"]
SECTION_NAME_CANDIDATES = ["section_name", "断面名称", "名称", "name", "Name", "河名", "沟道", "河流"]
X_CANDIDATES = ["x", "X", "lon", "Lon", "LON", "经度", "longitude"]
Y_CANDIDATES = ["y", "Y", "lat", "Lat", "LAT", "纬度", "latitude"]
WKT_CANDIDATES = ["wkt", "WKT", "geometry", "Geometry", "geom", "GEOM"]
VECTOR_SUFFIXES = {".shp", ".geojson", ".json", ".gpkg", ".csv", ".xls", ".xlsx"}
COMMON_VILLAGE_SUFFIXES = ("村民委员会", "村委会", "行政村", "自然村", "社区", "村")


@dataclass
class RasterItem:
    path: Path
    order: int
    label: str
    parsed_time: datetime | None
    hour_value: float | None


def first_existing_field(columns: Iterable[str], candidates: list[str]) -> str | None:
    cols = list(columns)
    lower_map = {str(c).lower(): c for c in cols}
    for candidate in candidates:
        if candidate in cols:
            return candidate
        found = lower_map.get(candidate.lower())
        if found is not None:
            return found
    return None


def resolve_vector_path(path: Path, preferred_words: tuple[str, ...] | None = None) -> Path:
    if not path.is_dir():
        return path
    candidates = [p for p in path.rglob("*") if p.is_file() and p.suffix.lower() in VECTOR_SUFFIXES]
    if not candidates:
        raise FileNotFoundError(f"文件夹中没有找到可读取的 GIS/表格文件：{path}")
    preferred_words = preferred_words or ("村", "边界", "boundary", "village")
    preferred = [p for p in candidates if any(word.lower() in str(p).lower() for word in preferred_words)]
    return sorted(preferred or candidates, key=lambda p: (len(p.parts), str(p)))[0]


def _normalize_text(value) -> str:
    text = "" if pd.isna(value) else str(value)
    return re.sub(r"\s+", "", text).strip()


def _village_aliases(value: str) -> set[str]:
    base = _normalize_text(value)
    aliases = {base} if base else set()
    for suffix in COMMON_VILLAGE_SUFFIXES:
        if base.endswith(suffix) and len(base) > len(suffix):
            aliases.add(base[: -len(suffix)])
    return {item for item in aliases if item}


def text_value_match_mask(series: pd.Series, value: str) -> pd.Series:
    aliases = _village_aliases(value)
    if not aliases:
        return pd.Series(False, index=series.index)
    normalized = series.map(_normalize_text)
    mask = pd.Series(False, index=series.index)
    for alias in aliases:
        escaped = re.escape(alias)
        exact = normalized == alias
        contains_value = normalized.str.contains(escaped, na=False)
        contained_by_value = normalized.map(lambda item: bool(item) and len(item) >= 2 and item in alias)
        mask = mask | exact | contains_value | contained_by_value
    return mask


def field_list(columns: Iterable[str]) -> str:
    return "、".join(str(col) for col in columns if str(col) != "geometry")


def read_table_as_geodata(path: Path, crs: str | None = None) -> gpd.GeoDataFrame:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        df = pd.read_csv(path)
    elif suffix in {".xls", ".xlsx"}:
        df = pd.read_excel(path)
    else:
        raise ValueError(f"不支持的表格格式：{path}")

    wkt_field = first_existing_field(df.columns, WKT_CANDIDATES)
    if wkt_field:
        geometry = df[wkt_field].apply(lambda value: wkt.loads(str(value)) if pd.notna(value) else None)
        return gpd.GeoDataFrame(df, geometry=geometry, crs=crs)

    x_field = first_existing_field(df.columns, X_CANDIDATES)
    y_field = first_existing_field(df.columns, Y_CANDIDATES)
    if not x_field or not y_field:
        raise ValueError(f"表格缺少 WKT 或 X/Y 坐标字段：{path}")
    geometry = [Point(xy) for xy in zip(df[x_field], df[y_field])]
    return gpd.GeoDataFrame(df, geometry=geometry, crs=crs or "EPSG:4326")


def read_vector(
    path: Path,
    crs_if_missing: str | None = None,
    encoding: str | None = None,
    where: str | None = None,
    bbox: tuple[float, float, float, float] | None = None,
    preferred_words: tuple[str, ...] | None = None,
) -> gpd.GeoDataFrame:
    path = resolve_vector_path(path, preferred_words=preferred_words)
    suffix = path.suffix.lower()
    if suffix in {".csv", ".xls", ".xlsx"}:
        gdf = read_table_as_geodata(path, crs=crs_if_missing)
    elif suffix in {".shp", ".geojson", ".json", ".gpkg"}:
        kwargs = {}
        if encoding:
            kwargs["encoding"] = encoding
        if where:
            kwargs["where"] = where
        if bbox:
            kwargs["bbox"] = bbox
        try:
            gdf = gpd.read_file(path, **kwargs)
        except TypeError:
            fallback_kwargs = dict(kwargs)
            fallback_kwargs.pop("where", None)
            try:
                gdf = gpd.read_file(path, **fallback_kwargs)
            except TypeError:
                fallback_kwargs.pop("bbox", None)
                gdf = gpd.read_file(path, **fallback_kwargs)
            if where:
                field, value = parse_simple_where(where)
                if field in gdf.columns:
                    gdf = gdf[gdf[field].astype(str) == value]
    else:
        raise ValueError(f"不支持的矢量数据格式：{path}")

    if gdf.empty:
        return gdf
    if gdf.crs is None and crs_if_missing:
        gdf = gdf.set_crs(CRS.from_user_input(crs_if_missing))
    return gdf


def parse_simple_where(where: str) -> tuple[str, str]:
    match = re.search(r'"?([^"=]+)"?\s*=\s*\'([^\']*)\'', where)
    if not match:
        return "", ""
    return match.group(1), match.group(2)


def build_where(field: str, value: str) -> str:
    safe_field = str(field).replace('"', '""')
    safe_value = str(value).replace("'", "''")
    return f'"{safe_field}" = \'{safe_value}\''


def find_matching_field(gdf: gpd.GeoDataFrame, value: str, candidates: list[str]) -> str | None:
    if not value:
        return first_existing_field(gdf.columns, candidates)
    matched_candidates = []
    for field in candidates:
        if field not in gdf.columns:
            match = first_existing_field(gdf.columns, [field])
            if match is None:
                continue
            field = match
        if text_value_match_mask(gdf[field], value).any():
            return field
        matched_candidates.append(field)
    for field in gdf.columns:
        if field == "geometry" or field in matched_candidates:
            continue
        if text_value_match_mask(gdf[field], value).any():
            return field
    return None


def filter_by_text_field(gdf: gpd.GeoDataFrame, field: str, value: str) -> gpd.GeoDataFrame:
    matched_field = first_existing_field(gdf.columns, [field])
    if matched_field is None:
        raise ValueError(f"字段不存在：{field}")
    field = matched_field
    mask = text_value_match_mask(gdf[field], value)
    exact = gdf[mask]
    if not exact.empty:
        return exact.copy()
    return exact.copy()


def load_village_boundary(
    auxiliary_path: Path,
    village_name: str,
    field: str | None = None,
    encoding: str | None = "gbk",
    crs_if_missing: str | None = None,
) -> gpd.GeoDataFrame:
    boundary = read_vector(auxiliary_path, crs_if_missing=crs_if_missing, encoding=encoding)
    if boundary.crs is None:
        raise ValueError("辅助村庄边界数据缺少 CRS。")
    if field:
        village_field = first_existing_field(boundary.columns, [field])
        if village_field is None:
            raise ValueError(f"配置的辅助村名字段不存在：{field}。可用字段：{field_list(boundary.columns)}")
    else:
        village_field = find_matching_field(boundary, village_name, VILLAGE_FIELD_CANDIDATES)
    if not village_field:
        raise ValueError(
            "辅助村庄边界中无法识别村名字段。"
            f"可用字段：{field_list(boundary.columns)}。"
            "请在高级参数中把“辅助村名字段”填写为真实村名字段，例如当前项目村界通常是 Name。"
        )
    selected = filter_by_text_field(boundary, village_field, village_name)
    if selected.empty:
        sample_values = boundary[village_field].dropna().astype(str).head(10).tolist()
        raise ValueError(f"辅助村庄边界字段 {village_field} 中找不到村庄：{village_name}。前几个值：{'、'.join(sample_values)}")
    return selected


def load_river_network(config: dict, logger) -> gpd.GeoDataFrame | None:
    river_path_text = str(config.get("river_network_path") or "").strip()
    if not river_path_text:
        return None
    river_path = Path(river_path_text)
    if not river_path.exists():
        raise FileNotFoundError(f"河流水系数据不存在：{river_path}")
    rivers = read_vector(
        river_path,
        crs_if_missing=config.get("river_network_crs"),
        encoding=config.get("river_network_encoding"),
        preferred_words=("河", "水系", "river", "stream", "沟", "渠"),
    )
    if rivers.empty:
        logger.warning("河流水系数据为空，动画将不显示河流。")
        return None
    if rivers.crs is None:
        raise ValueError("河流水系数据缺少 CRS，请在配置中填写 river_network_crs。")
    logger.info(f"河流水系读取成功：{len(rivers)} 条；CRS：{rivers.crs}")
    return rivers


def load_roads(config: dict, logger) -> gpd.GeoDataFrame | None:
    road_path_text = str(config.get("road_path") or "").strip()
    if not road_path_text:
        return None
    road_path = Path(road_path_text)
    if not road_path.exists():
        raise FileNotFoundError(f"道路/路网数据不存在：{road_path}")
    roads = read_vector(
        road_path,
        crs_if_missing=config.get("road_crs"),
        encoding=config.get("road_encoding"),
        preferred_words=("路", "道路", "公路", "road", "street", "transport"),
    )
    if roads.empty:
        logger.warning("道路/路网数据为空，将不按道路交点裁剪断面。")
        return None
    if roads.crs is None:
        raise ValueError("道路/路网数据缺少 CRS，请在配置中填写 road_crs。")
    logger.info(f"道路/路网读取成功：{len(roads)} 条；CRS：{roads.crs}")
    return roads


def load_buildings_for_village(config: dict, logger) -> tuple[gpd.GeoDataFrame, str | None, str]:
    building_path = Path(config["building_path"])
    if not building_path.exists():
        raise FileNotFoundError(f"村庄建筑物数据不存在：{building_path}")

    pre_field = str(config.get("building_prefilter_field") or "").strip()
    pre_value = str(config.get("building_prefilter_value") or "").strip()
    where = build_where(pre_field, pre_value) if pre_field and pre_value else None
    if where:
        logger.info(f"按属性预筛建筑物：{pre_field} = {pre_value}")

    buildings = read_vector(
        building_path,
        crs_if_missing=config.get("building_crs"),
        encoding=config.get("building_encoding"),
        where=where,
    )
    if buildings.empty:
        raise ValueError("建筑物数据为空，或预筛条件没有匹配任何建筑物。")
    if buildings.crs is None:
        raise ValueError("建筑物数据缺少 CRS，请为建筑物数据定义坐标系。")

    village_name = str(config.get("village_name") or "").strip()
    village_field = str(config.get("building_village_field") or "").strip() or None
    filter_method = "all_buildings"

    if village_name:
        if village_field:
            logger.info(f"使用配置字段筛选村庄建筑物：{village_field} = {village_name}")
            buildings = filter_by_text_field(buildings, village_field, village_name)
            filter_method = f"building_field:{village_field}"
        else:
            auto_field = find_matching_field(buildings, village_name, VILLAGE_FIELD_CANDIDATES)
            if auto_field:
                logger.info(f"自动识别建筑物村名字段：{auto_field}")
                buildings = filter_by_text_field(buildings, auto_field, village_name)
                village_field = auto_field
                filter_method = f"building_field:{auto_field}"
            elif config.get("auxiliary_path"):
                logger.info("建筑物中未识别到可用村名字段，改用辅助村庄边界裁剪建筑物。")
                boundary = load_village_boundary(
                    Path(config["auxiliary_path"]),
                    village_name,
                    field=config.get("auxiliary_village_field") or None,
                    encoding=config.get("auxiliary_encoding") or "gbk",
                    crs_if_missing=config.get("auxiliary_crs"),
                )
                boundary_for_buildings = boundary.to_crs(buildings.crs)
                minx, miny, maxx, maxy = boundary_for_buildings.total_bounds
                candidates = buildings.cx[minx:maxx, miny:maxy].copy()
                if candidates.empty:
                    buildings = candidates
                else:
                    mask = candidates.intersects(boundary_for_buildings.geometry.union_all())
                    buildings = candidates[mask].copy()
                filter_method = "auxiliary_boundary"
            else:
                raise ValueError("无法自动识别建筑物村名字段，也没有提供辅助村庄边界。")

    if buildings.empty:
        raise ValueError(f"筛选后没有建筑物：{village_name}")
    logger.info(f"参与分析建筑物数量：{len(buildings)}")
    return buildings, village_field, filter_method


def parse_raster_time(path: Path, fallback_order: int) -> tuple[str, datetime | None, float | None]:
    name = path.stem
    date_match = re.search(r"(20\d{6})[_-]?(\d{2})(\d{2})", name)
    if date_match:
        dt = datetime.strptime("".join(date_match.groups()), "%Y%m%d%H%M")
        return dt.strftime("%Y-%m-%d %H:%M"), dt, None

    hour_match = re.search(r"(?:^|[_-])H[_-]?0*(\d{1,3})(?:[_-]?00)?$", name, re.IGNORECASE)
    if not hour_match:
        hour_match = re.search(r"(?:^|[_-])0*(\d{1,3})h(?:$|[_-])", name, re.IGNORECASE)
    if not hour_match:
        hour_match = re.search(r"(?:^|[_-])t0*(\d{1,3})(?:$|[_-])", name, re.IGNORECASE)
    if hour_match:
        hour = float(hour_match.group(1))
        if hour.is_integer():
            return f"t{int(hour)}", None, hour
        return f"t{hour:g}", None, hour

    return f"t{fallback_order}", None, float(fallback_order)


def scan_flood_rasters(folder: Path, value_type: str, logger) -> list[RasterItem]:
    if not folder.exists():
        raise FileNotFoundError(f"模拟淹没数据文件夹不存在：{folder}")
    all_files = [p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in {".tif", ".tiff", ".out", ".asc"}]
    if not all_files:
        raise FileNotFoundError(f"没有找到 tif/out/asc 淹没栅格：{folder}")

    h_files = [p for p in all_files if re.match(r"^H_0*\d+_00$", p.stem, re.IGNORECASE)]
    if h_files:
        files = h_files
        logger.info(f"检测到 H_*.out 小时水深文件，优先使用 {len(files)} 个 H 文件。")
    else:
        files = all_files
        logger.info(f"扫描到 {len(files)} 个可用栅格文件。")

    items: list[RasterItem] = []
    for idx, path in enumerate(sorted(files), start=1):
        label, dt, hour = parse_raster_time(path, idx)
        items.append(RasterItem(path=path, order=idx, label=label, parsed_time=dt, hour_value=hour))

    items.sort(key=lambda item: (item.parsed_time is None, item.parsed_time or datetime.min, item.hour_value or item.order, item.path.name))
    for order, item in enumerate(items, start=1):
        item.order = order

    logger.info(f"淹没时间序列数量：{len(items)}；第一个时刻：{items[0].label}；最后一个时刻：{items[-1].label}")
    return items


def parse_section_dat(path: Path) -> list[tuple[float, float, float]]:
    points: list[tuple[float, float, float]] = []
    text = path.read_text(errors="ignore")
    for line in text.splitlines():
        parts = re.split(r"[，,\s]+", line.strip())
        nums: list[float] = []
        for item in parts:
            if not item:
                continue
            try:
                nums.append(float(item))
            except ValueError:
                continue
        for i in range(max(0, len(nums) - 2)):
            x, y, z = nums[i], nums[i + 1], nums[i + 2]
            if 300000 <= x <= 700000 and 4300000 <= y <= 4600000:
                points.append((x, y, z))
                break
    return points


def read_sections_from_dat_folder(folder: Path, section_crs: str, logger) -> gpd.GeoDataFrame:
    rows = []
    seen: set[str] = set()
    for path in folder.rglob("*"):
        if not path.is_file() or path.suffix.lower() != ".dat":
            continue
        key = str(path.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        if not re.match(r"^\d+$", path.stem):
            continue
        points = parse_section_dat(path)
        if len(points) < 2:
            continue
        parent_name = path.parent.name.lower()
        river_folder = path.parent.parent.name if parent_name == "dat" else path.parent.name
        rows.append(
            {
                "section_id": f"{river_folder}_{path.stem}",
                "section_name": f"{river_folder}_{path.stem}",
                "river_folder": river_folder,
                "source_path": str(path),
                "point_count": len(points),
                "geometry": LineString([(x, y) for x, y, _ in points]),
            }
        )
    if not rows:
        raise ValueError(f"没有从 DAT 文件生成任何断面线：{folder}")
    logger.info(f"从 DAT 文件生成断面线数量：{len(rows)}")
    return gpd.GeoDataFrame(rows, geometry="geometry", crs=CRS.from_user_input(section_crs))


def parse_section_paths(config: dict) -> list[Path]:
    raw_paths = config.get("section_paths")
    if raw_paths in (None, ""):
        raw_paths = config.get("section_path")
    if raw_paths in (None, ""):
        return []
    if isinstance(raw_paths, (list, tuple)):
        items = raw_paths
    else:
        text = str(raw_paths)
        items = [item.strip() for item in re.split(r"[;\n]+", text) if item.strip()]
    return [Path(str(item)) for item in items if str(item).strip()]


def _section_number_aliases(value: str) -> set[str]:
    text = _normalize_text(value)
    numbers = re.findall(r"\d+", text)
    aliases = {text} if text else set()
    for number in numbers:
        stripped = number.lstrip("0") or "0"
        aliases.update({number, stripped, f"{int(stripped):02d}", f"横断{int(stripped):02d}", f"横断{stripped}"})
    return {item for item in aliases if item}


def _find_section_summary_tables(config: dict, section_paths: list[Path]) -> list[Path]:
    raw = config.get("section_table_paths") or config.get("section_table_path")
    paths: list[Path] = []
    if raw not in (None, ""):
        items = raw if isinstance(raw, (list, tuple)) else [item.strip() for item in re.split(r"[;\n]+", str(raw)) if item.strip()]
        paths.extend(Path(str(item)) for item in items)
    else:
        for section_path in section_paths:
            roots = [section_path] if section_path.is_dir() else [section_path.parent]
            for root in roots:
                for item in root.rglob("*.xlsx"):
                    if item.name.startswith("~$"):
                        continue
                    if any(word in item.name for word in ("汇总", "断面数据", "横断")):
                        paths.append(item)
    seen = set()
    unique = []
    for path in paths:
        key = str(path.resolve()).lower() if path.exists() else str(path).lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def _make_section_trim_range(rows: list[dict]) -> dict | None:
    if not rows:
        return None
    rows = sorted(rows, key=lambda row: row["distance_m"])
    distances = [row["distance_m"] for row in rows]
    road_rows = []
    settlement_rows = []
    river_rows_all = []
    river_center_rows = []
    for row in rows:
        note = row.get("note", "")
        if ("路" in note or "街" in note) and "河道" not in note:
            road_rows.append(row)
        if any(word in note for word in ("村", "小区", "社区")):
            settlement_rows.append(row)
        if any(word in note for word in ("河道", "河中", "河底", "渠", "沟")):
            river_rows_all.append(row)
        if "河中" in note:
            river_center_rows.append(row)
    if not river_center_rows:
        return None
    river_center_rows = sorted(river_center_rows, key=lambda row: row["distance_m"])
    river_center_row = river_center_rows[len(river_center_rows) // 2]
    river_m = river_center_row["distance_m"]

    def choose_boundary(candidates: list[dict], side: str, label: str) -> tuple[dict | None, str | None]:
        if side == "before":
            before = [row for row in candidates if row["distance_m"] < river_m]
            return (max(before, key=lambda row: row["distance_m"]), label) if before else (None, None)
        after = [row for row in candidates if row["distance_m"] > river_m]
        return (min(after, key=lambda row: row["distance_m"]), label) if after else (None, None)

    start_row, start_label = choose_boundary(road_rows, "before", "路")
    if start_row is None:
        start_row, start_label = choose_boundary(settlement_rows, "before", "村庄/小区")
    if start_row is None:
        start_row = rows[0]
        start_label = "断面起点"

    end_row, end_label = choose_boundary(road_rows, "after", "路")
    if end_row is None:
        end_row, end_label = choose_boundary(settlement_rows, "after", "村庄/小区")
    if end_row is None:
        end_row = rows[-1]
        end_label = "断面末端"

    start_m = start_row["distance_m"]
    end_m = end_row["distance_m"]
    if end_m <= start_m:
        return None
    if not (start_m <= river_m <= end_m):
        return None
    trim_method = f"汇总表{start_label}-河中-{end_label}"
    segment_rows = [row for row in rows if start_m <= row["distance_m"] <= end_m and row.get("h85") is not None]
    river_rows = [
        row
        for row in segment_rows
        if any(word in row.get("note", "") for word in ("河道", "河中", "河底", "渠", "沟"))
    ]
    riverbed_elevation = None
    boundary_elevation = None
    original_depth = None
    if segment_rows and river_rows:
        center_rows = [row for row in river_rows if "河中" in row.get("note", "")]
        riverbed_elevation = min(row["h85"] for row in center_rows or river_rows)
        if start_row.get("h85") is not None and end_row.get("h85") is not None:
            boundary_elevation = min(start_row["h85"], end_row["h85"])
            original_depth = max(0.0, boundary_elevation - riverbed_elevation)
    return {
        "start_m": float(start_m),
        "end_m": float(end_m),
        "river_m": float(river_m),
        "road_count": len(road_rows),
        "settlement_count": len(settlement_rows),
        "river_count": len(river_rows_all),
        "section_table_method": trim_method,
        "section_original_depth_m": original_depth,
        "section_riverbed_elevation_m": riverbed_elevation,
        "section_boundary_elevation_m": boundary_elevation,
        "river_center_x": river_center_row.get("x"),
        "river_center_y": river_center_row.get("y"),
        "section_start_label": start_label,
        "section_end_label": end_label,
    }


def _extract_xy_from_section_table_row(row) -> tuple[float | None, float | None]:
    nums: list[float] = []
    for value in row:
        if value is None:
            continue
        try:
            nums.append(float(value))
        except (TypeError, ValueError):
            continue
    for idx in range(len(nums) - 1):
        first = nums[idx]
        second = nums[idx + 1]
        if 300000 <= first <= 700000 and 4300000 <= second <= 4600000:
            return first, second
        if 4300000 <= first <= 4600000 and 300000 <= second <= 700000:
            return second, first
    xs = [value for value in nums if 300000 <= value <= 700000]
    ys = [value for value in nums if 4300000 <= value <= 4600000]
    if xs and ys:
        return xs[0], ys[0]
    return None, None


def load_section_table_trim_ranges(config: dict, logger) -> dict[str, dict]:
    section_paths = parse_section_paths(config)
    manual_tables = config.get("section_table_paths") not in (None, "") or config.get("section_table_path") not in (None, "")
    tables = _find_section_summary_tables(config, section_paths)
    if not tables:
        return {}
    allow_plain_alias = manual_tables or len(tables) == 1

    ranges: dict[str, dict] = {}
    for table_path in tables:
        if not table_path.exists():
            logger.warning(f"断面汇总表不存在，已跳过：{table_path}")
            continue
        if table_path.suffix.lower() != ".xlsx":
            logger.warning(f"暂不支持读取 xls 断面汇总表，已跳过：{table_path}")
            continue
        try:
            workbook = load_workbook(table_path, data_only=True, read_only=True)
        except Exception as exc:
            logger.warning(f"断面汇总表读取失败，已跳过：{table_path}；原因：{exc}")
            continue
        worksheet = workbook.worksheets[0]
        by_section: dict[str, list[dict]] = {}
        for row in worksheet.iter_rows(min_row=4, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                section_name = str(row[1]).strip()
                note = "" if row[2] is None else str(row[2]).strip()
                distance_m = float(row[3])
                h85 = float(row[4]) if row[4] is not None else None
                x, y = _extract_xy_from_section_table_row(row)
            except Exception:
                continue
            if not section_name:
                continue
            by_section.setdefault(section_name, []).append(
                {"note": note, "distance_m": distance_m, "h85": h85, "x": x, "y": y}
            )
        loaded_count = 0
        table_river_folder = table_path.parent.name
        for section_name, rows in by_section.items():
            trim_range = _make_section_trim_range(rows)
            if trim_range is None:
                continue
            trim_range["section_table_path"] = str(table_path)
            trim_range["section_table_section"] = section_name
            trim_range["section_table_river_folder"] = table_river_folder
            for alias in _section_number_aliases(section_name):
                ranges[f"{table_river_folder}|{alias}"] = trim_range
                if allow_plain_alias:
                    ranges[alias] = trim_range
            loaded_count += 1
        logger.info(f"断面汇总表读取成功：{table_path.name}；可用于裁剪的断面 {loaded_count} 个。")
    return ranges


def normalize_section_fields(
    sections: gpd.GeoDataFrame,
    source_path: Path,
    config: dict,
    fallback_prefix: str,
) -> gpd.GeoDataFrame:
    sections = sections.copy()
    configured_id = str(config.get("section_id_field") or "").strip()
    configured_name = str(config.get("section_name_field") or "").strip()
    id_field = configured_id if configured_id in sections.columns else first_existing_field(sections.columns, SECTION_ID_CANDIDATES)
    name_field = configured_name if configured_name in sections.columns else first_existing_field(sections.columns, SECTION_NAME_CANDIDATES)

    if id_field:
        sections["section_id"] = sections[id_field].astype(str)
    elif "section_id" not in sections.columns:
        sections["section_id"] = [f"{fallback_prefix}_{i + 1}" for i in range(len(sections))]

    if name_field:
        sections["section_name"] = sections[name_field].astype(str)
    elif "section_name" not in sections.columns:
        sections["section_name"] = sections["section_id"].astype(str)

    if "source_path" not in sections.columns:
        sections["source_path"] = str(source_path)
    return sections


def load_sections(config: dict, logger) -> tuple[gpd.GeoDataFrame, str | None, str | None]:
    section_paths = parse_section_paths(config)
    if not section_paths:
        raise ValueError("没有选择河道断面数据。")
    section_crs = config.get("section_crs") or config.get("target_projected_crs") or config.get("raster_crs")
    loaded: list[gpd.GeoDataFrame] = []
    target_crs = CRS.from_user_input(section_crs) if section_crs else None

    for section_path in section_paths:
        if not section_path.exists():
            raise FileNotFoundError(f"河道断面数据不存在：{section_path}")
        if section_path.is_dir():
            one = read_sections_from_dat_folder(section_path, section_crs or "EPSG:4548", logger)
        else:
            one = read_vector(section_path, crs_if_missing=section_crs, encoding=config.get("section_encoding"))
            if one.crs is None and section_crs:
                one = one.set_crs(CRS.from_user_input(section_crs))
        if one.empty:
            logger.warning(f"断面数据为空，已跳过：{section_path}")
            continue
        if one.crs is None:
            raise ValueError(f"断面数据缺少 CRS，请填写 section_crs：{section_path}")
        if target_crs is None:
            target_crs = CRS.from_user_input(one.crs)
        elif one.crs != target_crs:
            one = one.to_crs(target_crs)
        one = normalize_section_fields(one, section_path, config, section_path.stem or "section")
        loaded.append(one)

    if not loaded:
        raise ValueError("没有读取到任何有效断面。")
    sections = gpd.GeoDataFrame(pd.concat(loaded, ignore_index=True), geometry="geometry", crs=target_crs)
    if sections.empty:
        raise ValueError("断面数据为空。")
    if sections.crs is None:
        raise ValueError("断面数据缺少 CRS，请填写 section_crs。")

    logger.info(f"断面数据路径数量：{len(section_paths)}；合并后断面数量：{len(sections)}")
    return sections, "section_id", "section_name"


